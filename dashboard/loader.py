"""
Загрузчик выгрузок 1С в базу. Использует ту же логику, что проверенные парсеры.
Идемпотентность: файл с тем же хешем не грузится повторно;
новая выгрузка того же года ЗАМЕНЯЕТ старые факты (перезаливающая загрузка).
"""
import re
import hashlib
from datetime import datetime
import openpyxl
from django.db import transaction
from .models import (Upload, SalesFact, SkuFact, DebtFact, DebtLine,
                     DebtSnapshot, DebtClientSnapshot, PackagingItem, Client)

MONTHS = ['янв', 'фев', 'мар', 'апр', 'май', 'июн', 'июл', 'авг', 'сен', 'окт', 'ноя', 'дек']
STM = ['вкусвилл', 'самокат', 'старс', 'fancy', 'butman', 'зеленая линия', 'зелёная линия',
       'true', 'бионова', 'ригла', 'dermadrop', 'sport club', 'армения', 'молдова', 'дубай', 'на арабском']
NONPROD = ['шоу-бокс', 'набор', 'дисплей', 'п/ф', 'полуфабрикат', 'напиток']


def _num(x):
    if x in (None, ''):
        return None
    if isinstance(x, (int, float)):
        return float(x)
    try:
        return float(str(x).replace('\xa0', '').replace(' ', '').replace(',', '.'))
    except ValueError:
        return None


def _date(x):
    if isinstance(x, datetime):
        return x.date()
    m = re.match(r'(\d{2})\.(\d{2})\.(\d{4})', str(x or ''))
    return datetime(int(m.group(3)), int(m.group(2)), int(m.group(1))).date() if m else None


def _hash(fileobj):
    fileobj.seek(0)
    h = hashlib.sha256(fileobj.read()).hexdigest()
    fileobj.seek(0)
    return h


def _read_rows(fileobj, filename=''):
    """Читает первый лист как список строк. Поддерживает и .xlsx (openpyxl), и старый .xls (xlrd)."""
    import io
    fileobj.seek(0)
    data = fileobj.read()
    fileobj.seek(0)
    name = (filename or getattr(fileobj, 'name', '') or '').lower()
    if name.endswith('.xls') and not name.endswith('.xlsx'):
        import xlrd
        book = xlrd.open_workbook(file_contents=data)
        sh = book.sheet_by_index(0)
        rows = []
        for r in range(sh.nrows):
            row = []
            for c in range(sh.ncols):
                cell = sh.cell(r, c)
                if cell.ctype == 3:  # дата в формате Excel
                    row.append(datetime(*xlrd.xldate_as_tuple(cell.value, book.datemode)))
                else:
                    row.append(cell.value if cell.value != '' else None)
            rows.append(tuple(row))
        return rows
    # .xlsx
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    ws = wb['Лист_1'] if 'Лист_1' in wb.sheetnames else wb[wb.sheetnames[0]]
    return list(ws.iter_rows(values_only=True))


def _month_cols(header):
    cols, year = {}, None
    for j, c in enumerate(header):
        if not c:
            continue
        m = re.match(r'([а-я]{3})[а-я]*\.?\s*(\d{2})', str(c).lower())
        if m and m.group(1) in MONTHS:
            cols[MONTHS.index(m.group(1)) + 1] = j
            year = 2000 + int(m.group(2))
    return cols, year


def _classify(name):
    low = name.lower()
    if any(k in low for k in NONPROD):
        return 'non_product'
    if any(k in low for k in STM):
        return 'private_label'
    return 'own'


@transaction.atomic
def load_sales_client(fileobj, filename, user=None):
    h = _hash(fileobj)
    if Upload.objects.filter(kind='sales_client', file_hash=h).exists():
        return {'skipped': True, 'reason': 'Такой файл уже загружен (совпал хеш)'}
    rows = _read_rows(fileobj, filename)
    cols, year = _month_cols(rows[0])
    up = Upload.objects.create(kind='sales_client', filename=filename, file_hash=h,
                               period_year=year, uploaded_by=user)
    # перезаливка: удаляем прошлые факты этого года
    SalesFact.objects.filter(year=year).delete()
    facts, total = [], 0
    for r in rows[2:]:
        doc = r[0]
        if doc is None or str(doc).strip() in ('', 'Итого') or not r[2]:
            continue
        s = str(doc).strip()
        dt = ('Реализация' if s.startswith('Реализа') else 'Корректировка' if s.startswith('Корректировк')
              else 'Комиссионер' if s.startswith('Отчет комисс') else 'Прочее')
        for mnum, col in cols.items():
            amt = _num(r[col + 1])
            qty = _num(r[col])
            if amt is None and qty is None:
                continue
            facts.append(SalesFact(upload=up, doc_type=dt, doc_date=_date(r[1]),
                                   client=str(r[2]).strip(), manager=(str(r[3]).strip() if r[3] else ''),
                                   year=year, month=mnum, qty=qty or 0, amount=int(amt or 0)))
            total += int(amt or 0)
    SalesFact.objects.bulk_create(facts, batch_size=1000)
    up.rows_loaded = len(facts)
    up.control_sum = total
    up.save()
    return {'skipped': False, 'year': year, 'rows': len(facts), 'total': total}


@transaction.atomic
def load_sales_sku(fileobj, filename, user=None):
    h = _hash(fileobj)
    if Upload.objects.filter(kind='sales_sku', file_hash=h).exists():
        return {'skipped': True, 'reason': 'Такой файл уже загружен'}
    rows = _read_rows(fileobj, filename)
    cols, year = _month_cols(rows[0])
    up = Upload.objects.create(kind='sales_sku', filename=filename, file_hash=h,
                               period_year=year, uploaded_by=user)
    SkuFact.objects.filter(year=year).delete()
    facts, total = [], 0
    for r in rows[2:]:
        name = r[0]
        if name is None or str(name).strip() in ('', 'Итого'):
            continue
        name = str(name).strip()
        bt = _classify(name)
        for mnum, col in cols.items():
            amt = _num(r[col + 1])
            qty = _num(r[col])
            if amt is None and qty is None:
                continue
            facts.append(SkuFact(upload=up, sku_raw=name, brand_type=bt, year=year,
                                 month=mnum, qty=qty or 0, amount=int(amt or 0)))
            total += int(amt or 0)
    SkuFact.objects.bulk_create(facts, batch_size=1000)
    up.rows_loaded = len(facts)
    up.control_sum = total
    up.save()
    return {'skipped': False, 'year': year, 'rows': len(facts), 'total': total}


# корзины срока просрочки в новом отчёте (индексы колонок H..M)
_DEBT_BUCKETS = [(7, 'До 7 дн'), (8, '8–15 дн'), (9, '16–30 дн'),
                 (10, '31–40 дн'), (11, '41–90 дн'), (12, '>90 дн')]


def _parse_debt_new(rows, filename):
    """Новый формат «по срокам долга»: строка клиента + под ней реализации.
    Колонки: 0 Менеджер/Документ · 1 Покупатель · 2 Отгрузка · 3 Срок оплаты ·
    4 Долг · 5 Просрочено · 6 Дней · 7..12 корзины срока. Возвращает СЛОВАРИ (без записи в БД)."""
    m = re.search(r'на (\d{2})\.(\d{2})\.(\d{4})', filename)
    snap = datetime(int(m.group(3)), int(m.group(2)), int(m.group(1))).date() if m else None
    hdr = next((i for i, r in enumerate(rows)
                if len(r) > 1 and r[1] and str(r[1]).strip() == 'Покупатель'), None)
    if hdr is None:
        return [], [], 0
    facts, lines, total, cur = [], [], 0, None
    for r in rows[hdr + 2:]:
        g = lambda i: r[i] if i < len(r) else None    # безопасный доступ (строки бывают короче)
        a = str(r[0]).strip() if r and r[0] else ''
        b = str(g(1)).strip() if g(1) else ''
        tot = _num(g(4))
        if b and b != 'Итого':                # строка клиента (Покупатель в колонке B)
            cur = b
            if tot:
                facts.append(dict(client=cur, manager=a, snapshot_date=snap,
                                  ship_date=_date(g(2)), due_date=_date(g(3)),
                                  debt_total=int(tot), debt_overdue=int(_num(g(5)) or 0),
                                  overdue_days=int(_num(g(6)) or 0)))
                total += int(tot)
        elif a and not b and cur and tot and a != 'Итого':   # любой документ под клиентом
            bucket = next((nm for idx, nm in _DEBT_BUCKETS if _num(g(idx))), '')
            mno = re.search(r'№\s*(\S+)', a)
            lines.append(dict(client=cur, doc_no=(mno.group(1) if mno else ''),
                              ship_date=_date(g(2)), due_date=_date(g(3)),
                              debt_total=int(tot), debt_overdue=int(_num(g(5)) or 0),
                              overdue_days=int(_num(g(6)) or 0), bucket=bucket))
    # дедуп фантомных повторов: строка без номера, дублирующая ту же строку С номером
    # (клиент+отгрузка+срок+сумма) — убираем, чтобы не задваивать график оплат/кошельки
    numbered = {(l['client'], l['ship_date'], l['due_date'], l['debt_total'])
                for l in lines if l['doc_no']}
    seen, dedup = set(), []
    for l in lines:
        k = (l['client'], l['ship_date'], l['due_date'], l['debt_total'], l['doc_no'])
        if not l['doc_no'] and k[:4] in numbered:
            continue                         # безномерной дубль строки с номером
        if k in seen:
            continue                         # полный дубль (та же строка дважды)
        seen.add(k)
        dedup.append(l)
    return facts, dedup, total


def _parse_debt_old(rows):
    """Старый широкий формат: «Покупатель» в 1-й колонке. Возвращает словари."""
    hdr = next((i for i, r in enumerate(rows) if r[0] and str(r[0]).strip() == 'Покупатель'), None)
    if hdr is None:
        return [], [], 0
    facts, total = [], 0
    for r in rows[hdr + 1:]:
        g = lambda i: r[i] if i < len(r) else None
        client = r[0] if r else None
        if client is None or str(client).strip() in ('', 'Итого'):
            continue
        tot = _num(g(16))
        if tot is None:
            continue
        facts.append(dict(client=str(client).strip(),
                          manager=(str(g(5)).strip() if g(5) else ''),
                          ship_date=_date(g(11)), due_date=_date(g(13)),
                          debt_total=int(tot), debt_overdue=int(_num(g(18)) or 0),
                          overdue_days=int(_num(g(21)) or 0)))
        total += int(tot)
    return facts, [], total


@transaction.atomic
def load_debt(fileobj, filename, user=None):
    h = _hash(fileobj)
    if Upload.objects.filter(kind='debt', file_hash=h).exists():
        return {'skipped': True, 'reason': 'Такой файл уже загружен'}
    rows = _read_rows(fileobj, filename)
    # строгое распознавание нового формата: в шапке «Менеджер…» в кол.A и «Покупатель» в кол.B
    is_new = any(len(r) > 1 and r[1] and str(r[1]).strip() == 'Покупатель'
                 and r[0] and 'енеджер' in str(r[0]) for r in rows[:8])
    if is_new:
        facts, lines, total = _parse_debt_new(rows, filename)
    else:
        facts, lines, total = _parse_debt_old(rows)
    # ВАЖНО: не трогаем существующую дебиторку, пока разбор не дал непустой результат,
    # иначе кривой/чужой файл обнулит данные
    if not facts:
        return {'skipped': True, 'reason': 'Формат не распознан или нет строк — данные не изменены'}
    snap_date = facts[0].get('snapshot_date') or datetime.now().date()
    up = Upload.objects.create(kind='debt', filename=filename, file_hash=h, uploaded_by=user,
                               note=f'снимок {snap_date}')
    # СНИМОК ПО ДАТЕ: заменяем только этот же день, прошлые снимки храним
    DebtFact.objects.filter(snapshot_date=snap_date).delete()
    DebtLine.objects.filter(snapshot_date=snap_date).delete()
    DebtFact.objects.bulk_create(
        [DebtFact(upload=up, **{**d, 'snapshot_date': snap_date}) for d in facts], batch_size=1000)
    DebtLine.objects.bulk_create(
        [DebtLine(upload=up, snapshot_date=snap_date, **d) for d in lines], batch_size=1000)
    up.rows_loaded = len(facts)
    up.control_sum = total
    up.save()
    # сводка снимка для истории/динамики
    overdue = sum(f['debt_overdue'] for f in facts)
    DebtSnapshot.objects.update_or_create(
        date=snap_date, defaults={'total': total, 'overdue': overdue, 'count': len(facts)})
    DebtClientSnapshot.objects.filter(date=snap_date).delete()
    DebtClientSnapshot.objects.bulk_create(
        [DebtClientSnapshot(date=snap_date, client=f['client'],
                            debt_total=f['debt_total'], debt_overdue=f['debt_overdue']) for f in facts],
        batch_size=1000)
    # ретеншен: храним последние 120 снимков
    old = list(DebtSnapshot.objects.order_by('-date').values_list('date', flat=True)[120:])
    if old:
        DebtFact.objects.filter(snapshot_date__in=old).delete()
        DebtLine.objects.filter(snapshot_date__in=old).delete()
        DebtSnapshot.objects.filter(date__in=old).delete()
        DebtClientSnapshot.objects.filter(date__in=old).delete()
    return {'skipped': False, 'rows': len(facts), 'lines': len(lines), 'total': total, 'snap': str(snap_date)}


# ---------- Снимок упаковки (обновление остатков с сайта) ----------
def pack_key(name):
    """Нормализованный ключ имени из снимка упаковки — для сопоставления с справочником."""
    s = str(name or '').replace('\xa0', ' ').strip().lower()
    s = re.sub(r'^\d+\s*', '', s)
    s = s.replace('упаковка', '').replace('"', '')
    s = re.sub(r'\(.*?\)', '', s)
    s = re.sub(r'[^а-яёa-z0-9 ]', ' ', s)
    return re.sub(r'\s+', ' ', s).strip()


def load_packaging_snapshot(fileobj, filename, user=None):
    """Обновляет остатки упаковки из последнего листа файла-снимка.
    Остаток может стоять в той же строке или со смещением на строку ниже (грязный снимок)."""
    fileobj.seek(0)
    try:
        wb = openpyxl.load_workbook(fileobj, data_only=True)
    except Exception:
        return {'skipped': True, 'reason': 'Не удалось открыть файл (нужен .xlsx)'}
    ws = wb.worksheets[-1]                       # последний лист = самый свежий снимок
    rows = list(ws.iter_rows(values_only=True))
    stock_by_key = {}
    for i, r in enumerate(rows):
        a = r[0] if len(r) > 0 else None
        b = r[1] if len(r) > 1 else None
        if not a:
            continue
        val = b
        if val in (None, 0) and i + 1 < len(rows):
            a2 = rows[i + 1][0] if len(rows[i + 1]) > 0 else None
            b2 = rows[i + 1][1] if len(rows[i + 1]) > 1 else None
            if (a2 is None or str(a2).strip() == '') and b2 not in (None, 0):
                val = b2
        num = _num(val)
        if num is None:
            continue
        k = pack_key(a)
        if k:
            stock_by_key[k] = int(num)
    updated, missed = 0, []
    for it in PackagingItem.objects.all():
        key = it.snap_key or pack_key(it.upak)
        if key in stock_by_key:
            it.stock = stock_by_key[key]
            it.save(update_fields=['stock'])
            updated += 1
        else:
            missed.append(it.upak)
    Upload.objects.create(kind='packaging', filename=filename, file_hash=_hash(fileobj),
                          uploaded_by=user, rows_loaded=updated, note=f'лист {ws.title}')
    return {'skipped': False, 'sheet': ws.title, 'updated': updated,
            'total': PackagingItem.objects.count(), 'missed': missed}


# ---------- Справочник контрагентов из 1С (.mxl) ----------
def _parse_mxl(fileobj):
    """Разбор родного формата 1С (MOXCEL) в список строк {индекс_колонки: значение}."""
    fileobj.seek(0)
    data = fileobj.read()
    if isinstance(data, bytes):
        data = data.decode('utf-8', 'ignore')
    cr = re.compile(r'\{16,\d+,\s*\{1,(?:1,\s*\{"#","((?:[^"\\]|\\.)*)"\}\s*|0)\},0\},(\d+),')
    rows, cur, last = [], {}, 0
    for m in cr.finditer(data):
        v, c = m.group(1), int(m.group(2))
        if c <= last and cur:
            rows.append(cur)
            cur = {}
        if v is not None:
            cur[c] = v.strip()
        last = c
    if cur:
        rows.append(cur)
    return rows


def _mxl_cols(rows):
    """Индексы колонок по строке-заголовку."""
    hdr = {}
    for r in rows:
        if 'ИНН' in r.values() and 'Полное наименование' in r.values():
            hdr = {v: k for k, v in r.items()}
            break
    col = lambda n: next((hdr[k] for k in hdr if k.startswith(n)), None)
    return {'name': col('Наименование в программе'), 'inn': col('ИНН'),
            'full': col('Полное наименование'), 'code': col('Код'),
            'phone': col('Телефон'), 'contact': col('Контактное лицо'),
            'email': col('Email'), 'manager': col('Менеджер'),
            'city': col('Город'), 'addr': col('Юридический адрес')}


_GOV_RE = re.compile(r'ИФНС|ФНС|ОСФР|ПФР|\bФСС\b|КАЗНАЧ|НАЛОГОВ|МИНИСТ|АДМИНИСТРАЦ|ПЕНСИОН|'
                     r'\bБАНК\b|\bКБ\b|МОДУЛЬБАНК|ТОЧКА БАНК|ТИНЬКОФ|СБЕРБАНК|\bФОНД\b|УФК\b', re.I)
_FIO_RE = re.compile(r'^[А-ЯЁ][А-ЯЁа-яё-]+\s+[А-ЯЁ][А-ЯЁа-яё-]+\s+[А-ЯЁ][А-ЯЁа-яё-]+$')


def _is_junk(name, full):
    if not name:
        return True
    if _GOV_RE.search(name + ' ' + full):
        return True
    if _FIO_RE.match(name) and 'ИП' not in full.upper():   # физлицо-учредитель, не ИП
        return True
    return False


@transaction.atomic
def load_contractors(combined, suppliers, user=None):
    """Клиенты = общий справочник МИНУС поставщики, без гос/банков/учредителей.
    Обновляет реквизиты в справочнике клиентов, не трогая канал/лимит/исключение/заметку."""
    from datetime import date
    comb = _parse_mxl(combined)
    c = _mxl_cols(comb)
    if not c['inn'] or not c['name']:
        return {'skipped': True, 'reason': 'Не распознан формат .mxl (нет колонок ИНН/Наименование)'}
    sup_rows = _parse_mxl(suppliers)
    sc = _mxl_cols(sup_rows)
    sup_inns = {r.get(sc['inn']) for r in sup_rows
                if re.fullmatch(r'\d{10}|\d{12}', str(r.get(sc['inn'], '')))}
    created = updated = skipped_junk = 0
    today = date.today()
    for r in comb:
        inn = str(r.get(c['inn'], ''))
        if not re.fullmatch(r'\d{10}|\d{12}', inn):
            continue
        if inn in sup_inns:
            continue
        name = r.get(c['name'], '').strip()
        full = r.get(c['full'], '').strip()
        if _is_junk(name, full):
            skipped_junk += 1
            continue
        obj, is_new = Client.objects.get_or_create(name=name)
        obj.inn = inn
        if full and not obj.full_name:
            obj.full_name = full
        # реквизиты из 1С только ЗАПОЛНЯЮТ пустое — ручные правки менеджера не затираем
        for field, col in (('manager', 'manager'), ('phone', 'phone'), ('contact', 'contact'),
                           ('email', 'email'), ('city', 'city'), ('address', 'addr')):
            val = (r.get(c[col]) or '').strip()
            if val and not getattr(obj, field):
                setattr(obj, field, val)
        obj.synced_at = today
        obj.save()
        created += is_new
        updated += not is_new
    Upload.objects.create(kind='contractors', filename='справочник контрагентов',
                          file_hash=_hash(combined), uploaded_by=user, rows_loaded=created + updated)
    return {'skipped': False, 'created': created, 'updated': updated,
            'junk': skipped_junk, 'suppliers': len(sup_inns)}
