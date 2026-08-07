from datetime import datetime, date, timedelta
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.contrib.auth.views import LoginView
from . import metrics, loader
from .models import Upload

MONTHS = ['янв', 'фев', 'мар', 'апр', 'май', 'июн', 'июл', 'авг', 'сен', 'окт', 'ноя', 'дек']
MONTHS_FULL = ['январь', 'февраль', 'март', 'апрель', 'май', 'июнь', 'июль',
               'август', 'сентябрь', 'октябрь', 'ноябрь', 'декабрь']
CUR_YEAR, PREV_YEAR = 2026, 2025


def _pdate(s):
    try:
        return datetime.strptime(s, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return None


def _filters(request):
    return {'manager': request.GET.get('manager') or None,
            'channel': request.GET.get('channel') or None,
            'client': request.GET.get('client') or None,
            'date_from': _pdate(request.GET.get('date_from')),
            'date_to': _pdate(request.GET.get('date_to'))}


def _period_presets():
    """Быстрые периоды: неделя / месяц / квартал / год (относительно сегодня)."""
    t = date.today()
    q_start = date(t.year, ((t.month - 1) // 3) * 3 + 1, 1)
    iso = lambda d: d.strftime('%Y-%m-%d')
    return [
        {'label': 'Неделя', 'df': iso(t - timedelta(days=6)), 'dt': iso(t)},
        {'label': 'Месяц', 'df': iso(date(t.year, t.month, 1)), 'dt': iso(t)},
        {'label': 'Квартал', 'df': iso(q_start), 'dt': iso(t)},
        {'label': 'Год', 'df': iso(date(t.year, 1, 1)), 'dt': iso(t)},
    ]


def _base_ctx(request, page):
    f = _filters(request)
    opts = metrics.filter_options(CUR_YEAR)
    last_sales = Upload.objects.filter(kind='sales_client').order_by('-uploaded_at').first()
    last_debt = Upload.objects.filter(kind='debt').order_by('-uploaded_at').first()
    return {'page': page, 'cur_year': CUR_YEAR, 'prev_year': PREV_YEAR, 'f': f, 'opts': opts,
            'last_sales': last_sales, 'last_debt': last_debt,
            'months': MONTHS,
            'sel_manager': f['manager'] or '', 'sel_channel': f['channel'] or '',
            'sel_client': f['client'] or '',
            'sel_from': request.GET.get('date_from', ''), 'sel_to': request.GET.get('date_to', ''),
            'has_period': bool(f['date_from'] and f['date_to']),
            'presets': _period_presets()}


@login_required
def svodka(request):
    c = _base_ctx(request, 'svodka')
    f = c['f']
    year = None if c['has_period'] else CUR_YEAR
    s = metrics.sales_summary(year, **f)
    years = metrics.sales_years()
    try:
        chart_y = int(request.GET.get('y'))
    except (TypeError, ValueError):
        chart_y = None
    if chart_y not in years:
        chart_y = CUR_YEAR if CUR_YEAR in years or not years else years[0]
    y = metrics.yoy(chart_y, chart_y - 1, **{k: f[k] for k in ('manager', 'channel', 'client')})
    ymax = max(max(y['now']), max(y['prev'])) or 1
    now7, prev7 = sum(y['now'][:7]), sum(y['prev'][:7]) or 1
    day = metrics.sales_by_day(year, **f)
    dmax = max([d['amount'] for d in day['days']], default=1) or 1
    d = metrics.debt_summary(manager=f['manager'], client=f['client'])
    c.update({
        'sales_total': s['total'], 'returns': abs(s['returns']),
        'yoy_bars': [{'m': MONTHS[i], 'now_h': round(y['now'][i] / ymax * 100),
                      'prev_h': round(y['prev'][i] / ymax * 100),
                      'now_v': y['now'][i], 'prev_v': y['prev'][i]} for i in range(12)],
        'yoy_delta': round((now7 - prev7) / prev7 * 100),
        'chart_y': chart_y, 'chart_prev': chart_y - 1, 'sales_years': years,
        'day': day, 'day_bars': [{'d': x['day'], 'h': round(x['amount'] / dmax * 100),
                                  'amount': x['amount']} for x in day['days']],
        'top_clients': metrics.top_clients(year, 5, **f),
        'managers': metrics.by_manager(year, **f),
        'debt': d, 'debtors': d['debtors'][:4],
        'plans': metrics.plan_status(year, manager=f['manager'],
                                     date_from=f['date_from'], date_to=f['date_to']),
    })
    return render(request, 'dashboard/svodka.html', c)


@login_required
def prodazhi(request):
    c = _base_ctx(request, 'prodazhi')
    f = c['f']
    year = None if c['has_period'] else CUR_YEAR
    s = metrics.sales_summary(year, **f)
    y = metrics.yoy(CUR_YEAR, PREV_YEAR, **{k: f[k] for k in ('manager', 'channel', 'client')})
    now7, prev7 = sum(y['now'][:7]), sum(y['prev'][:7]) or 1
    c.update({
        'sales_total': s['total'], 'returns': abs(s['returns']),
        'yoy_delta': round((now7 - prev7) / prev7 * 100),
        'all_clients': metrics.all_clients(year, **f),
        'managers': metrics.by_manager(year, **f),
        'top_sku': metrics.top_sku(CUR_YEAR, 8),
        'plans': metrics.plan_status(year, manager=f['manager'],
                                     date_from=f['date_from'], date_to=f['date_to']),
    })
    return render(request, 'dashboard/prodazhi.html', c)


@login_required
def debitorka(request):
    c = _base_ctx(request, 'debitorka')
    f = c['f']
    only_overdue = request.GET.get('overdue') == '1'
    order = request.GET.get('order', '-debt_total')
    snap_dates = metrics.debt_dates()
    sel_snap = _pdate(request.GET.get('snap'))
    if sel_snap not in snap_dates:
        sel_snap = snap_dates[0] if snap_dates else None
    d = metrics.debt_summary(manager=f['manager'], client=f['client'],
                             only_overdue=only_overdue, order=order, snapshot=sel_snap)
    aging = metrics.debt_aging(manager=f['manager'], client=f['client'], snapshot=sel_snap)
    bmax = max([b['amount'] for b in aging['buckets']], default=1) or 1
    for b in aging['buckets']:
        b['h'] = round(b['amount'] / bmax * 100)
    hist = metrics.debt_history()
    hmax = max([h['total'] for h in hist], default=1) or 1
    hist_bars = [{'date': h['date'], 'total': h['total'], 'overdue': h['overdue'],
                  'h': round(h['total'] / hmax * 100)} for h in hist]
    c.update({'debt': d, 'debtors': d['debtors'], 'aging': aging,
              'hist_bars': hist_bars, 'only_overdue': only_overdue, 'order': order,
              'snap_dates': snap_dates, 'sel_snap': sel_snap, 'is_latest': sel_snap == (snap_dates[0] if snap_dates else None)})
    return render(request, 'dashboard/debitorka.html', c)


@login_required
def debtor(request, client):
    c = _base_ctx(request, 'debitorka')
    d = metrics.debt_summary(client=client)
    lines = metrics.debt_lines(client)          # реальная расшифровка по реализациям из 1С
    fallback = []
    if not lines:                                # расшифровки нет — старый приблизительный список
        fallback = metrics.client_sales(client)
        for x in fallback:
            x['date_str'] = x['doc_date'].strftime('%d.%m.%Y') if x['doc_date'] else '—'
    c.update({'client_name': client, 'debt': d,
              'debtor_row': d['debtors'][0] if d['debtors'] else None,
              'lines': lines, 'sales': fallback})
    return render(request, 'dashboard/debtor.html', c)


@login_required
def upload(request):
    if not request.user.is_staff:                 # загрузка данных — только для админов
        return render(request, 'dashboard/upload.html',
                      {'page': 'upload', 'msg': [], 'no_access': True})
    msg = []
    if request.method == 'POST':
        force = bool(request.POST.get('force'))
        fn_map = {'sales_client': loader.load_sales_client,
                  'sales_sku': loader.load_sales_sku, 'debt': loader.load_debt}
        for key, fn in fn_map.items():
            f = request.FILES.get(key)
            if f:
                r = fn(f, f.name, request.user, force=force)
                msg.append(f'{key}: пропущено — {r["reason"]}' if r.get('skipped')
                           else f'{key}: загружено {r["rows"]} строк, сумма {r["total"]:,} ₽')
        fp = request.FILES.get('packaging')
        if fp:
            r = loader.load_packaging_snapshot(fp, fp.name, request.user)
            if r.get('skipped'):
                msg.append(f'упаковка: пропущено — {r["reason"]}')
            else:
                m = f' · не сопоставлено: {len(r["missed"])}' if r['missed'] else ''
                msg.append(f'упаковка: обновлено остатков {r["updated"]} из {r["total"]} (лист {r["sheet"]}){m}')
        ca, cs = request.FILES.get('contractors_all'), request.FILES.get('contractors_sup')
        if ca and cs:
            r = loader.load_contractors(ca, cs, request.user)
            if r.get('skipped'):
                msg.append(f'контрагенты: пропущено — {r["reason"]}')
            else:
                msg.append(f'контрагенты: добавлено {r["created"]}, обновлено {r["updated"]}, '
                           f'отсеяно мусора {r["junk"]} (поставщиков {r["suppliers"]})')
        elif ca or cs:
            msg.append('контрагенты: нужны ОБА файла — общий и поставщики')
    return render(request, 'dashboard/upload.html', {'msg': msg, 'page': 'upload'})


@login_required
def export_xlsx(request):
    import io
    import openpyxl
    from django.http import HttpResponse
    kind = request.GET.get('kind', 'clients')
    wb = openpyxl.Workbook()
    ws = wb.active
    if kind == 'debt':
        ws.title = 'Дебиторка'
        d = metrics.debt_summary()
        ws.append(['Клиент', 'Долг', 'Просрочено', 'Дней просрочки', 'Срок оплаты', 'Менеджер'])
        for x in d['debtors']:
            ws.append([x['client'], x['debt_total'], x['debt_overdue'], x['overdue_days'],
                       x['due_date'].strftime('%d.%m.%Y') if x['due_date'] else '', x['manager']])
    elif kind == 'rfm':
        ws.title = 'Сегменты'
        data = metrics.rfm()
        ws.append(['Клиент', 'Сегмент', 'ABC', 'Последний заказ', 'Дней', 'Заказов', 'Выручка'])
        for r in data['rows']:
            ws.append([r['client'], r['seg'], r['abc'], r['last'].strftime('%d.%m.%Y'),
                       r['r_days'], r['f'], r['m']])
    else:
        kind = 'clients'
        ws.title = 'Клиенты'
        years, rows = metrics.clients_list(CUR_YEAR)
        ws.append(['Клиент', 'ИНН', 'Канал'] + [f'Продажи {y}' for y in years] + ['Долг', 'Просрочено'])
        for r in rows:
            ws.append([r['name'], r['inn'], r['channel']] + r['by_year'] + [r['debt'], r['overdue']])
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    resp = HttpResponse(bio.read(),
                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{kind}.xlsx"'
    return resp


@login_required
def managers(request):
    return render(request, 'dashboard/managers.html',
                  {'page': 'managers', 'rows': metrics.managers_list(CUR_YEAR), 'cur_year': CUR_YEAR})


@login_required
def manager_card(request, manager):
    p = metrics.manager_profile(manager)
    ymax = max([b['total'] for b in p['by_year']], default=1) or 1
    for b in p['by_year']:
        b['h'] = round(b['total'] / ymax * 100)
    return render(request, 'dashboard/manager_card.html', {'page': 'managers', 'p': p})


@login_required
def rfm(request):
    data = metrics.rfm()
    seg = request.GET.get('seg') or ''
    abc = request.GET.get('abc') or ''
    rows = data['rows']
    if seg:
        rows = [r for r in rows if r['seg'] == seg]
    if abc:
        rows = [r for r in rows if r['abc'] == abc]
    rows = sorted(rows, key=lambda r: -r['m'])[:400]
    return render(request, 'dashboard/rfm.html', {
        'page': 'rfm', 'segments': data['segments'], 'rows': rows,
        'total_clients': data['total_clients'], 'sel_seg': seg, 'sel_abc': abc,
        'shown': len(rows)})


@login_required
def cost(request):
    from .models import CostGroup
    groups = list(CostGroup.objects.all())
    try:
        gid = int(request.GET.get('group'))
    except (TypeError, ValueError):
        gid = None
    sel = next((g for g in groups if g.id == gid), None)
    sel_mgr = request.GET.get('manager') or ''
    managers = metrics.filter_options(CUR_YEAR)['managers']
    ctx = {'page': 'cost', 'groups': groups, 'sel_group': sel,
           'managers': managers, 'sel_mgr': sel_mgr,
           'last_sales': Upload.objects.filter(kind='sales_client').order_by('-uploaded_at').first(),
           'last_debt': Upload.objects.filter(kind='debt').order_by('-uploaded_at').first()}
    empty = {'mapped': [], 'unmapped': [], 'vat_pct': 22, 'avg_margin': 0}
    if sel_mgr:
        ctx['report'] = metrics.manager_report(sel_mgr)
        ctx['report_label'] = 'менеджер ' + sel_mgr
        ctx['c'] = empty
    elif sel and sel.clients.exists():
        ctx['report'] = metrics.group_report(list(sel.clients.values_list('name', flat=True)))
        ctx['report_label'] = 'группа «' + sel.name + '»'
        ctx['c'] = empty
    else:
        ctx['c'] = metrics.cost_margin(group=sel)
    return render(request, 'dashboard/cost.html', ctx)


@login_required
def cost_channel(request, code):
    return render(request, 'dashboard/cost_channel.html',
                  {'page': 'cost', 'ch': metrics.channel_positions(code)})


@login_required
def cost_map(request):
    from django.db.models import Max
    from django.shortcuts import redirect
    from .models import CostItem, CostSku, SkuFact
    if not request.user.is_staff:
        return render(request, 'dashboard/cost_map.html', {'page': 'cost', 'no_access': True})
    if request.method == 'POST':
        cid = request.POST.get('cost')
        act = request.POST.get('action')
        if act == 'add' and request.POST.get('sku'):
            sku = request.POST['sku'].strip()
            if SkuFact.objects.filter(sku_raw=sku).exists():   # только реальный SKU
                CostSku.objects.get_or_create(cost_id=cid, sku=sku)
        elif act == 'remove_extra':
            CostSku.objects.filter(id=request.POST.get('csid')).delete()
        elif act == 'remove_primary':
            CostItem.objects.filter(id=cid).update(sku='')
        return redirect(f"{request.path}#c{cid}")
    ly = SkuFact.objects.aggregate(y=Max('year'))['y']
    skus = sorted(set(SkuFact.objects.filter(year=ly).values_list('sku_raw', flat=True))) if ly else []
    items = []
    for it in CostItem.objects.prefetch_related('skus').order_by('line', 'name'):
        att = []
        if it.sku:
            att.append({'label': it.sku, 'primary': True})
        for x in it.skus.all():
            att.append({'label': x.sku, 'primary': False, 'csid': x.id})
        items.append({'id': it.id, 'name': it.name, 'line': it.line, 'cost': round(it.cost), 'att': att})
    return render(request, 'dashboard/cost_map.html', {'page': 'cost', 'items': items, 'skus': skus,
                  'mapped': sum(1 for i in items if i['att']), 'total': len(items)})


@login_required
def signals(request):
    return render(request, 'dashboard/signals.html',
                  {'page': 'signals', 's': metrics.signals(), 'cur_year': CUR_YEAR})


@login_required
def sravnenie(request):
    import json
    ov = metrics.year_overview()
    chart_json = '{}'
    palette = ['#9aa0aa', '#3b82f6', '#6d5bd0', '#2fa84f', '#f59e0b']
    if ov.get('years'):
        for i, r in enumerate(ov['revenue']):
            pass
        chart = {
            'years': [str(y) for y in ov['years']],
            'months': MONTHS,
            'colors': {str(y): palette[i % len(palette)] for i, y in enumerate(ov['years'])},
            'monthly': {str(y): [round(v) for v in ov['monthly'][y]] for y in ov['years']},
        }
        chart_json = json.dumps(chart, ensure_ascii=False)
        rmax = max((r['total'] for r in ov['revenue']), default=1) or 1
        for r in ov['revenue']:
            r['h'] = round(r['total'] / rmax * 100)
    forgotten_n = len(metrics.forgotten_clients(CUR_YEAR)) if ov.get('years') else 0
    return render(request, 'dashboard/sravnenie.html', {'page': 'sravnenie', 'ov': ov,
                  'months': MONTHS, 'forgotten_n': forgotten_n, 'cur_year': CUR_YEAR,
                  'chart_json': chart_json})


@login_required
def sravnenie_sku(request):
    q = request.GET.get('q', '')
    data = metrics.sku_year_detail(q)
    return render(request, 'dashboard/sravnenie_detail.html', {
        'page': 'sravnenie', 'kind': 'sku', 'title': 'Все SKU по годам',
        'years': data['years'], 'rows': data['rows'][:400], 'q': q, 'total_rows': len(data['rows'])})


@login_required
def sravnenie_clients(request):
    q = request.GET.get('q', '')
    data = metrics.client_year_detail(q)
    return render(request, 'dashboard/sravnenie_detail.html', {
        'page': 'sravnenie', 'kind': 'client', 'title': 'Все клиенты по годам',
        'years': data['years'], 'rows': data['rows'][:400], 'q': q, 'total_rows': len(data['rows'])})


@login_required
def sravnenie_drill(request):
    kind = request.GET.get('kind')
    try:
        year = int(request.GET.get('year'))
    except (TypeError, ValueError):
        year = CUR_YEAR
    if kind == 'forgotten':
        rows = metrics.forgotten_clients(year)
        title = f'Забытые клиенты'
        sub = f'покупали раньше, но в {year} — ни разу'
    elif kind == 'lost':
        rows = metrics.lost_clients(year)
        title = f'Клиенты, ушедшие в {year}'
        sub = f'покупали в {year - 1}, но не в {year}'
    else:
        lo = int(request.GET.get('lo') or 0)
        hi = request.GET.get('hi')
        hi = int(hi) if hi else None
        rows = metrics.bucket_clients(year, lo, hi)
        rng = f'{lo // 1000}к–{hi // 1000}к' if hi else f'{lo // 1000}к+'
        absent = request.GET.get('absent') == '1'
        if absent:                       # только те, кто в текущем году не заказывал
            cur_set = set(metrics._year_client_sales(CUR_YEAR))
            rows = [r for r in rows if r['name'] not in cur_set]
        title = f'Клиенты {rng} ₽ · {year}' + (f' · не заказывали в {CUR_YEAR}' if absent else '')
        sub = (f'из корзины {rng} за {year}, кто в {CUR_YEAR} не сделал ни заказа'
               if absent else f'выручка за {year} в диапазоне {rng} ₽')
        return render(request, 'dashboard/sravnenie_drill.html', {
            'page': 'sravnenie', 'rows': rows, 'title': title, 'sub': sub,
            'total': sum(r['sales'] for r in rows), 'year': year, 'bucket': True,
            'lo': lo, 'hi': hi or '', 'absent': absent, 'cur_year': CUR_YEAR,
            'show_absent': year != CUR_YEAR})
    return render(request, 'dashboard/sravnenie_drill.html', {
        'page': 'sravnenie', 'rows': rows, 'title': title, 'sub': sub,
        'total': sum(r['sales'] for r in rows), 'year': year})


@login_required
def clients(request):
    q = (request.GET.get('q') or '').strip().lower()
    years, all_rows = metrics.clients_list(CUR_YEAR)
    rows = [r for r in all_rows if q in r['name'].lower() or q in r['inn']] if q else all_rows
    last_sales = Upload.objects.filter(kind='sales_client').order_by('-uploaded_at').first()
    last_debt = Upload.objects.filter(kind='debt').order_by('-uploaded_at').first()
    return render(request, 'dashboard/clients.html', {
        'page': 'clients', 'rows': rows, 'years': years, 'q': request.GET.get('q', ''), 'found': len(rows),
        'total_clients': len(all_rows), 'cur_year': CUR_YEAR,
        'total_sales': sum(r['sales'] for r in rows), 'total_debt': sum(r['debt'] for r in rows),
        'last_sales': last_sales, 'last_debt': last_debt})


@login_required
def client_card(request, client):
    from .models import Client, ManagerProfile
    years = metrics.sales_years()
    try:
        sel_y = int(request.GET.get('y'))
    except (TypeError, ValueError):
        sel_y = None
    if sel_y not in years:
        sel_y = years[0] if years else CUR_YEAR
    p = metrics.client_profile(client, sel_y, sel_y - 1)
    prof = ManagerProfile.objects.filter(user=request.user).first()
    can_edit = bool(request.user.is_staff or (prof and (prof.can_edit_all
                    or (prof.manager and prof.manager == p['owner_manager']))))
    saved = False
    if request.method == 'POST' and request.POST.get('action') == 'delete' and request.user.is_staff:
        from django.shortcuts import redirect
        Client.objects.filter(name=client).delete()
        return redirect('clients')
    if request.method == 'POST' and can_edit:
        obj, _ = Client.objects.get_or_create(name=client)
        for f in ('channel', 'status', 'full_name', 'kpp', 'ogrn', 'phone', 'contact', 'email',
                  'city', 'address', 'bank', 'account', 'bik', 'payment_terms', 'retro_bonus',
                  'contract', 'delivery', 'note'):
            setattr(obj, f, request.POST.get(f, '').strip())
        lim = request.POST.get('credit_limit', '').replace(' ', '')
        obj.credit_limit = int(lim) if lim.isdigit() else None
        mo = request.POST.get('min_order', '').replace(' ', '')
        obj.min_order = int(mo) if mo.isdigit() else None
        obj.save()
        p = metrics.client_profile(client, CUR_YEAR, PREV_YEAR)
        saved = True
    last_sales = Upload.objects.filter(kind='sales_client').order_by('-uploaded_at').first()
    last_debt = Upload.objects.filter(kind='debt').order_by('-uploaded_at').first()
    ymax = max(max(p['now']), max(p['prev'])) or 1
    bars = [{'m': MONTHS[i], 'now_h': round(p['now'][i] / ymax * 100),
             'prev_h': round(p['prev'][i] / ymax * 100),
             'now_v': p['now'][i], 'prev_v': p['prev'][i]} for i in range(12)]
    hmax = max([h['debt_total'] for h in p['hist']], default=1) or 1
    hist = [{'date': h['date'], 'total': h['debt_total'], 'h': round(h['debt_total'] / hmax * 100)} for h in p['hist']]
    return render(request, 'dashboard/client_card.html', {
        'page': 'clients', 'client_name': client, 'p': p, 'bars': bars, 'hist': hist,
        'cur_year': sel_y, 'prev_year': sel_y - 1, 'sales_years': years, 'sel_y': sel_y,
        'can_edit': can_edit, 'saved': saved, 'channels': Client.CHANNELS, 'statuses': Client.STATUS,
        'last_sales': last_sales, 'last_debt': last_debt})


@login_required
def oplaty(request):
    ym = request.GET.get('ym', '')
    try:
        y, m = map(int, ym.split('-'))
        date(y, m, 1)
    except (ValueError, TypeError):
        t = date.today()
        y, m = t.year, t.month
    cal = metrics.payment_calendar(y, m)
    prev = date(y, m, 1) - timedelta(days=1)
    nxt = date(y + (m == 12), (m % 12) + 1, 1)
    return render(request, 'dashboard/oplaty.html', {
        'page': 'oplaty', 'cal': cal, 'month_name': MONTHS_FULL[m - 1], 'year': y,
        'prev_ym': f'{prev.year}-{prev.month:02d}', 'next_ym': f'{nxt.year}-{nxt.month:02d}',
        'weekdays': ['Пн', 'Вт', 'Ср', 'Чт', 'Пт', 'Сб', 'Вс'],
        'forecast': metrics.payment_forecast(8),
        'last_debt': Upload.objects.filter(kind='debt').order_by('-uploaded_at').first(),
    })


@login_required
def oplaty_day(request, day):
    try:
        d = datetime.strptime(day, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        d = date.today()
    rows = metrics.payment_day(d)
    return render(request, 'dashboard/oplaty_day.html', {
        'page': 'oplaty', 'day': d, 'rows': rows,
        'total': sum(r['debt_total'] for r in rows),
        'overdue': sum(r['debt_overdue'] for r in rows),
        'back_ym': f'{d.year}-{d.month:02d}',
    })


@login_required
def upakovka(request):
    series = request.GET.get('series') or None
    used = request.GET.get('used') or None
    rows = metrics.packaging_status(series=series, used=used)
    crit = [r for r in rows if r['status'] == 'crit']
    warn = [r for r in rows if r['status'] == 'warn']
    return render(request, 'dashboard/upakovka.html', {
        'page': 'upakovka', 'rows': rows, 'crit': crit, 'warn': warn,
        'series_list': metrics.packaging_series_list(),
        'sel_series': series or '', 'sel_used': used or '',
        'cur_year': CUR_YEAR, 'prev_year': PREV_YEAR,
    })


# ---------- Лиды / потенциальные клиенты ----------
_LEAD_HMAP = {
    'компания': 'company', 'название': 'company', 'наименование': 'company', 'company': 'company', 'контрагент': 'company',
    'инн': 'inn', 'inn': 'inn',
    'канал': 'channel', 'channel': 'channel',
    'город': 'city', 'city': 'city',
    'контакт': 'contact', 'контактное лицо': 'contact', 'фио': 'contact', 'contact': 'contact', 'лицо': 'contact',
    'телефон': 'phone', 'тел': 'phone', 'phone': 'phone', 'моб': 'phone',
    'email': 'email', 'почта': 'email', 'e-mail': 'email', 'мейл': 'email', 'емейл': 'email',
    'сайт': 'website', 'website': 'website', 'url': 'website',
    'источник': 'source', 'source': 'source', 'откуда': 'source',
    'соцсети': 'socials', 'соцсеть': 'socials', 'соц.сети': 'socials', 'socials': 'socials', 'соц': 'socials',
    'статус': 'status', 'status': 'status',
    'ответственный': 'owner', 'менеджер': 'owner', 'owner': 'owner',
    'потенциал': 'potential', 'сумма': 'potential', 'бюджет': 'potential', 'potential': 'potential', 'сделка': 'potential',
    'заметки': 'note', 'заметка': 'note', 'комментарий': 'note', 'note': 'note', 'коммент': 'note',
}


def _to_int(s):
    s = str(s or '').replace(' ', '').replace('\xa0', '').replace(',', '.')
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None


LEAD_ACTIONS = {
    'add': 'Добавлен лид', 'edit': 'Изменён лид', 'move': 'Перемещён по воронке',
    'convert': 'Заведён клиентом', 'delete': 'Удалён лид', 'note': 'Добавлена запись',
    'import': 'Импорт лидов', 'stages': 'Изменены этапы',
}


def _lead_log(request, action, lead=None, company='', detail=''):
    from .models import LeadLog
    try:
        LeadLog.objects.create(
            user=(request.user.get_username() if request.user.is_authenticated else ''),
            action=action, lead_id=(lead.id if lead else None),
            company=(company or (lead.company if lead else '')), detail=detail)
    except Exception:
        pass   # логирование не должно ломать основное действие


def _parse_leads_file(f):
    """Читает xlsx/csv с лидами → список dict. Заголовки распознаются гибко (рус/eng)."""
    from .models import Lead
    ch_map = {l.lower(): c for c, l in Lead.CHANNELS}
    ch_map.update({c: c for c, _ in Lead.CHANNELS})
    name = (getattr(f, 'name', '') or '').lower()
    if name.endswith('.csv'):
        import io
        import csv
        raw = f.read()
        text = None
        for enc in ('utf-8-sig', 'cp1251', 'utf-8'):
            try:
                text = raw.decode(enc)
                break
            except (UnicodeDecodeError, AttributeError):
                text = None
        if text is None:
            text = raw.decode('utf-8', 'ignore')
        delim = ';' if text.count(';') > text.count(',') else ','
        table = [list(r) for r in csv.reader(io.StringIO(text), delimiter=delim)]
    else:
        import io
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
        ws = wb.active
        table = [[('' if c is None else c) for c in row] for row in ws.iter_rows(values_only=True)]
    table = [r for r in table if any(str(c).strip() for c in r)]
    if not table:
        return []
    headers = [str(h or '').strip().lower() for h in table[0]]
    idx = {}
    for i, h in enumerate(headers):
        fld = _LEAD_HMAP.get(h)
        if fld and fld not in idx:
            idx[fld] = i
    if 'company' not in idx:               # нет шапки → первый столбец = компания
        idx = {'company': 0}
        data = table
    else:
        data = table[1:]
    out = []
    for row in data:
        def g(k):
            i = idx.get(k)
            return str(row[i]).strip() if (i is not None and i < len(row) and row[i] is not None) else ''
        rec = {k: g(k) for k in ('company', 'inn', 'city', 'contact', 'phone', 'email', 'website', 'socials', 'source', 'owner', 'note')}
        rec['channel'] = ch_map.get(g('channel').lower(), '')
        rec['stage_name'] = g('status')          # текст этапа из файла — сопоставим при импорте
        rec['potential'] = _to_int(g('potential'))
        if rec['company']:
            out.append(rec)
    return out


def _default_stage():
    from .models import LeadStage
    return LeadStage.objects.order_by('order', 'id').first()


def _stage_by_id(sid):
    from .models import LeadStage
    try:
        return LeadStage.objects.filter(id=int(sid)).first()
    except (TypeError, ValueError):
        return None


def _lead_add_from_post(request):
    from .models import Lead
    comp = (request.POST.get('company') or '').strip()
    if comp:
        stage = _stage_by_id(request.POST.get('stage')) or _default_stage()
        lead = Lead.objects.create(
            company=comp, inn=(request.POST.get('inn') or '').strip(),
            channel=request.POST.get('channel') or '', city=(request.POST.get('city') or '').strip(),
            contact=(request.POST.get('contact') or '').strip(), phone=(request.POST.get('phone') or '').strip(),
            email=(request.POST.get('email') or '').strip(), website=(request.POST.get('website') or '').strip(),
            socials=(request.POST.get('socials') or '').strip(),
            source=(request.POST.get('source') or '').strip(), owner=(request.POST.get('owner') or '').strip(),
            potential=_to_int(request.POST.get('potential')), stage=stage)
        _lead_log(request, 'add', lead, detail='этап: %s' % (stage.name if stage else '—'))


def _lead_filter(request):
    from .models import Lead
    sel_channel = request.GET.get('channel') or ''
    sel_owner = request.GET.get('owner') or ''
    q = (request.GET.get('q') or '').strip().lower()
    rows = list(Lead.objects.select_related('stage').all())
    if sel_channel:
        rows = [r for r in rows if r.channel == sel_channel]
    if sel_owner:
        rows = [r for r in rows if r.owner == sel_owner]
    if q:
        rows = [r for r in rows if q in r.company.lower() or q in (r.inn or '') or q in (r.city or '').lower()]
    return rows, sel_channel, sel_owner


@login_required
def leads(request):
    """Канбан-доска: колонки — настраиваемые этапы воронки, карточки перетаскиваются мышкой."""
    from django.shortcuts import redirect
    from .models import Lead, LeadStage
    if request.method == 'POST' and request.POST.get('action') == 'add':
        _lead_add_from_post(request)
        return redirect('leads')
    rows, sel_channel, sel_owner = _lead_filter(request)
    today = date.today()
    for r in rows:
        r.overdue = bool(r.next_action and r.next_action < today
                         and not (r.stage and (r.stage.is_won or r.stage.is_lost)))
    stages = list(LeadStage.objects.all())
    default_id = stages[0].id if stages else None
    cols = []
    for st in stages:
        items = [r for r in rows if r.stage_id == st.id or (r.stage_id is None and st.id == default_id)]
        cols.append({'id': st.id, 'name': st.name, 'color': st.color, 'is_won': st.is_won, 'is_lost': st.is_lost,
                     'items': items, 'n': len(items), 'sum': sum(r.potential or 0 for r in items)})
    from .models import SalesManager
    return render(request, 'dashboard/leads_board.html', {
        'page': 'leads', 'cols': cols, 'stages': stages, 'total': len(rows),
        'sel_channel': sel_channel, 'sel_owner': sel_owner, 'q': request.GET.get('q', ''), 'channels': Lead.CHANNELS,
        'managers': list(SalesManager.objects.filter(active=True))})


@login_required
def leads_list(request):
    """Табличный вид лидов с воронкой-фильтром по этапам."""
    from django.shortcuts import redirect
    from .models import Lead, LeadStage
    if request.method == 'POST' and request.POST.get('action') == 'add':
        _lead_add_from_post(request)
        return redirect('leads_list')
    stages = list(LeadStage.objects.all())
    all_rows = list(Lead.objects.select_related('stage').all())
    counts = {st.id: 0 for st in stages}
    for r in all_rows:
        if r.stage_id in counts:
            counts[r.stage_id] += 1
    funnel = [{'id': st.id, 'name': st.name, 'n': counts.get(st.id, 0)} for st in stages]
    sel_stage = request.GET.get('stage') or ''
    rows, sel_channel, sel_owner = _lead_filter(request)
    if sel_stage:
        rows = [r for r in rows if str(r.stage_id) == sel_stage]
    won = sum(1 for r in all_rows if r.stage and r.stage.is_won)
    lost = sum(1 for r in all_rows if r.stage and r.stage.is_lost)
    from .models import SalesManager
    return render(request, 'dashboard/leads.html', {
        'page': 'leads', 'rows': rows, 'funnel': funnel, 'total': len(all_rows), 'found': len(rows),
        'sel_stage': sel_stage, 'sel_channel': sel_channel, 'sel_owner': sel_owner, 'q': request.GET.get('q', ''),
        'channels': Lead.CHANNELS, 'stages': stages, 'managers': list(SalesManager.objects.filter(active=True)),
        'won': won, 'active': len(all_rows) - won - lost})


@login_required
def lead_move(request, lead_id):
    """Перемещение карточки между этапами (drag-and-drop)."""
    from django.http import JsonResponse
    from django.shortcuts import get_object_or_404
    from .models import Lead
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    stage = _stage_by_id(request.POST.get('stage'))
    if not stage:
        return JsonResponse({'ok': False, 'error': 'bad stage'}, status=400)
    lead = get_object_or_404(Lead, id=lead_id)
    old = lead.stage.name if lead.stage else '—'
    lead.stage = stage
    lead.save(update_fields=['stage', 'updated_at'])
    _lead_log(request, 'move', lead, detail='%s → %s' % (old, stage.name))
    return JsonResponse({'ok': True, 'id': lead.id, 'stage': stage.id})


@login_required
def lead_stages(request):
    """Настройка этапов воронки: добавить/переименовать/переставить/удалить."""
    from django.shortcuts import redirect
    from .models import LeadStage
    if request.method == 'POST':
        act = request.POST.get('action')
        if act == 'add':
            nm = (request.POST.get('name') or '').strip()
            if nm:
                mx = LeadStage.objects.count()
                LeadStage.objects.create(name=nm, order=mx)
        elif act == 'save':
            for st in LeadStage.objects.all():
                nm = (request.POST.get('name_%d' % st.id) or '').strip()
                od = _to_int(request.POST.get('order_%d' % st.id))
                cl = (request.POST.get('color_%d' % st.id) or '').strip()
                st.name = nm or st.name
                st.order = od if od is not None else st.order
                if cl:
                    st.color = cl
                st.is_won = request.POST.get('won_%d' % st.id) == 'on'
                st.is_lost = request.POST.get('lost_%d' % st.id) == 'on'
                st.save()
        elif act == 'delete':
            sid = _stage_by_id(request.POST.get('sid'))
            if sid and LeadStage.objects.count() > 1:
                sid.delete()
        _lead_log(request, 'stages', detail=act)
        return redirect('lead_stages')
    stages = list(LeadStage.objects.all())
    return render(request, 'dashboard/lead_stages.html', {'page': 'leads', 'stages': stages})


def _lead_erp(company):
    """Данные из 1С по компании: продажи по годам + текущий долг. None, если нет."""
    from .models import SalesFact, DebtClientSnapshot
    from django.db.models import Sum
    sales_rows = list(SalesFact.objects.filter(client=company)
                      .values('year').annotate(a=Sum('amount')).order_by('-year'))
    dsnap = DebtClientSnapshot.objects.filter(client=company).order_by('-date').first()
    mgr = (SalesFact.objects.filter(client=company).exclude(manager='')
           .values_list('manager', flat=True).first())
    if not (sales_rows or dsnap):
        return None
    return {
        'sales': [{'year': r['year'], 'amount': round(r['a'] or 0)} for r in sales_rows],
        'debt': round(dsnap.debt_total) if dsnap else 0,
        'overdue': round(dsnap.debt_overdue) if dsnap else 0,
        'debt_date': dsnap.date if dsnap else None,
        'manager': mgr,
    }


@login_required
def lead_delete(request, lead_id):
    """Удаление лида (для перетаскивания в «помойку»)."""
    from django.http import JsonResponse
    from django.shortcuts import get_object_or_404
    from .models import Lead
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    lead = get_object_or_404(Lead, id=lead_id)
    _lead_log(request, 'delete', company=lead.company, detail='ID %s' % lead.id)
    lead.delete()
    return JsonResponse({'ok': True})


@login_required
def lead_quick(request, lead_id):
    """Компактная карточка-превью лида для модалки на доске."""
    from django.shortcuts import get_object_or_404
    from .models import Lead, Client
    lead = get_object_or_404(Lead, id=lead_id)
    return render(request, 'dashboard/lead_quick.html', {
        'lead': lead, 'erp': _lead_erp(lead.company),
        'is_client': Client.objects.filter(name=lead.company).exists(),
        'notes': list(lead.notes.all()[:4])})


def _convert_lead_to_client(lead, user):
    """Закрытие цикла: завести лид клиентом в справочник."""
    from .models import Client, LeadStage, LeadNote
    obj, _created = Client.objects.get_or_create(name=lead.company.strip())
    if lead.inn and not obj.inn:
        obj.inn = lead.inn
    if lead.city and not obj.city:
        obj.city = lead.city
    if lead.phone and not obj.phone:
        obj.phone = lead.phone
    if lead.email and not obj.email:
        obj.email = lead.email
    if lead.contact and not obj.contact:
        obj.contact = lead.contact
    if lead.channel and not obj.channel:
        obj.channel = lead.channel
    obj.save()
    lead.converted = True
    won = LeadStage.objects.filter(is_won=True).order_by('order').first()
    if won:
        lead.stage = won
    lead.save()
    LeadNote.objects.create(lead=lead, text='✅ Заведён в справочник клиентов',
                            author=(user.get_username() if user else ''))
    return obj


@login_required
def lead_logs(request):
    """Журнал действий по лидам — только для админов."""
    from .models import LeadLog
    if not request.user.is_staff:
        return render(request, 'dashboard/lead_logs.html', {'page': 'leads', 'no_access': True})
    sel_user = request.GET.get('user') or ''
    sel_action = request.GET.get('action') or ''
    qs = LeadLog.objects.all()
    if sel_user:
        qs = qs.filter(user=sel_user)
    if sel_action:
        qs = qs.filter(action=sel_action)
    logs = list(qs[:500])
    for l in logs:
        l.action_label = LEAD_ACTIONS.get(l.action, l.action)
    users = sorted(set(LeadLog.objects.exclude(user='').values_list('user', flat=True)))
    return render(request, 'dashboard/lead_logs.html', {
        'page': 'leads', 'logs': logs, 'users': users, 'actions': LEAD_ACTIONS,
        'sel_user': sel_user, 'sel_action': sel_action, 'total': LeadLog.objects.count()})


@login_required
def lead_card(request, lead_id):
    from django.shortcuts import redirect, get_object_or_404
    from .models import Lead, LeadStage, LeadNote
    lead = get_object_or_404(Lead, id=lead_id)
    saved = converted = False
    if request.method == 'POST':
        act = request.POST.get('action')
        if act == 'delete':
            _lead_log(request, 'delete', company=lead.company, detail='ID %s' % lead.id)
            lead.delete()
            return redirect('leads')
        if act == 'note':
            txt = (request.POST.get('text') or '').strip()
            if txt:
                LeadNote.objects.create(lead=lead, text=txt, author=request.user.get_username())
                _lead_log(request, 'note', lead, detail=txt[:120])
            return redirect('lead_card', lead_id=lead.id)
        if act == 'convert':
            _convert_lead_to_client(lead, request.user)
            _lead_log(request, 'convert', lead)
            return redirect('lead_card', lead_id=lead.id)
        for fld in ('company', 'inn', 'channel', 'city', 'contact', 'phone', 'email',
                    'website', 'socials', 'source', 'owner', 'note'):
            setattr(lead, fld, (request.POST.get(fld) or '').strip())
        lead.stage = _stage_by_id(request.POST.get('stage')) or lead.stage
        lead.potential = _to_int(request.POST.get('potential'))
        lead.last_touch = _pdate((request.POST.get('last_touch') or '').strip())
        lead.next_action = _pdate((request.POST.get('next_action') or '').strip())
        lead.save()
        _lead_log(request, 'edit', lead, detail='этап: %s' % (lead.stage.name if lead.stage else '—'))
        saved = True
    from .models import Client, SalesManager
    is_client = Client.objects.filter(name=lead.company).exists()
    erp = _lead_erp(lead.company)   # 360°: данные из 1С, если компания есть в базе
    return render(request, 'dashboard/lead_card.html', {
        'page': 'leads', 'lead': lead, 'saved': saved, 'converted': converted,
        'notes': list(lead.notes.all()), 'is_client': is_client, 'erp': erp,
        'channels': Lead.CHANNELS, 'stages': list(LeadStage.objects.all()),
        'managers': list(SalesManager.objects.filter(active=True))})


@login_required
def lead_import(request):
    from .models import Lead, LeadStage
    result = None
    if request.method == 'POST' and request.FILES.get('file'):
        try:
            parsed = _parse_leads_file(request.FILES['file'])
        except Exception as e:
            return render(request, 'dashboard/lead_import.html', {'page': 'leads', 'error': str(e)})
        stage_by_name = {s.name.lower(): s for s in LeadStage.objects.all()}
        default_stage = _default_stage()
        existing_inn = set(x for x in Lead.objects.exclude(inn='').values_list('inn', flat=True))
        existing_names = {c.lower() for c in Lead.objects.values_list('company', flat=True)}
        created = skipped = 0
        for r in parsed:
            inn = r['inn']
            if (inn and inn in existing_inn) or r['company'].lower() in existing_names:
                skipped += 1
                continue
            stage = stage_by_name.get((r.get('stage_name') or '').lower(), default_stage)
            Lead.objects.create(
                company=r['company'], inn=inn, channel=r['channel'], city=r['city'],
                contact=r['contact'], phone=r['phone'], email=r['email'], website=r['website'],
                socials=r.get('socials', ''), source=r['source'] or 'импорт', owner=r['owner'], stage=stage,
                potential=r.get('potential'), note=r['note'])
            if inn:
                existing_inn.add(inn)
            existing_names.add(r['company'].lower())
            created += 1
        result = {'created': created, 'skipped': skipped, 'total': len(parsed)}
        _lead_log(request, 'import', detail='создано %s, пропущено %s' % (created, skipped))
    return render(request, 'dashboard/lead_import.html', {'page': 'leads', 'result': result})


class Login(LoginView):
    template_name = 'dashboard/login.html'


def logout_view(request):
    """Выход по обычной ссылке (GET). Встроенный LogoutView в Django 5 требует POST → 405."""
    from django.contrib.auth import logout
    from django.shortcuts import redirect
    logout(request)
    return redirect('login')
